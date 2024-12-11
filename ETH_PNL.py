
from dotenv import load_dotenv
import os
import pandas as pd
from dune_client.client import DuneClient
from dune_client.query import QueryBase
from dune_client.types import QueryParameter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class WalletReport:
    def __init__(self, wallet_address):
        self.wallet_address = wallet_address
        load_dotenv()
        self.dune_api_key = os.getenv('DUNE_API_KEY')
        self.request_timeout = int(os.getenv('DUNE_API_REQUEST_TIMEOUT'))
        self.dune = DuneClient(
            api_key=self.dune_api_key,
            base_url="https://api.dune.com",
            request_timeout=self.request_timeout
        )
        self.TRANSACTION_QUERY_ID = 3831623
        self.SUMMARY_QUERY_ID = 3831751
        self.parameters = [
            QueryParameter.text_type(name='day', value='-30'),
            QueryParameter.text_type(name='wallet', value=self.wallet_address)
        ]
        # self.output_file_path = f"{self.wallet_address}.xlsx"
        self.reports_folder = "reports"
        os.makedirs(self.reports_folder, exist_ok=True)
        self.output_file_path = os.path.join(self.reports_folder, f"{self.wallet_address}.xlsx")

        self.summary_df = None
        self.transaction_df = None

    def fetch_data(self):
        transaction_query = QueryBase(query_id=self.TRANSACTION_QUERY_ID, params=self.parameters)
        summary_query = QueryBase(query_id=self.SUMMARY_QUERY_ID, params=self.parameters)

        self.transaction_df = self.dune.run_query_dataframe(transaction_query, performance='medium')
        self.transaction_df.columns = [col.lower() for col in self.transaction_df.columns]

        self.summary_df = self.dune.run_query_dataframe(summary_query, performance='medium')
        self.summary_df.columns = [col.lower() for col in self.summary_df.columns]

    def save_to_excel(self):
        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            self.summary_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=0)
            pd.DataFrame([[]]).to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=len(self.summary_df) + 1)
            self.transaction_df.to_excel(writer, sheet_name='Sheet1', index=False, startrow=len(self.summary_df) + 2)

    def apply_formatting(self):
        workbook = load_workbook(self.output_file_path)
        worksheet = workbook.active

        brown_fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 1:
                    cell.font = Font(bold=True)
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        summary_row = 1
        total_spent_amount_col = None
        actual_profit_col = None
        win_rate_col = None
        pnl_r_col = None
        pnl_l_col = None
        loss_rate_col = None

        for cell in worksheet[summary_row]:
            if cell.value == 'total_spent_amount':
                total_spent_amount_col = cell.column
            elif cell.value == 'actual_profit':
                actual_profit_col = cell.column
            elif cell.value == 'win_rate':
                win_rate_col = cell.column
            elif cell.value == 'pnl_r':
                pnl_r_col = cell.column
            elif cell.value == 'pnl_l':
                pnl_l_col = cell.column
            elif cell.value == 'loss_rate':
                loss_rate_col = cell.column
        

        summary_cols = [total_spent_amount_col, actual_profit_col, win_rate_col, pnl_r_col, pnl_l_col, loss_rate_col]
        if None in summary_cols:
            raise ValueError("Required summary columns not found in the Excel sheet.")

        header_row = 4
        delta_eth_col = None
        delta_percentage_col = None
        dexscreener_col = None
        number_buys_col = None
        number_sells_col = None
        token_symbol_col = None
        outcome_col = None
        incoming_col = None

        for cell in worksheet[header_row]:
            if cell.value == 'delta_eth':
                delta_eth_col = cell.column
            elif cell.value == 'delta_percentage':
                delta_percentage_col = cell.column
            elif cell.value == 'dexscreener':
                dexscreener_col = cell.column
            elif cell.value == 'number_buys':
                number_buys_col = cell.column
            elif cell.value == 'number_sells':
                number_sells_col = cell.column
            elif cell.value == 'token_symbol':
                token_symbol_col = cell.column
            elif cell.value == 'outcome':
                outcome_col = cell.column
            elif cell.value == 'incoming':
                incoming_col = cell.column
        if dexscreener_col is not None:
            dexscreener_column_letter = get_column_letter(dexscreener_col)
            worksheet.column_dimensions[dexscreener_column_letter].width = 20

        required_cols = [delta_eth_col, delta_percentage_col, dexscreener_col, number_buys_col, number_sells_col, token_symbol_col, outcome_col, incoming_col]
        if None in required_cols:
            raise ValueError("Required columns not found in the Excel sheet.")

        total_spent_amount_cell = worksheet.cell(row=summary_row + 1, column=total_spent_amount_col)
        pnl_r_cell = worksheet.cell(row=summary_row + 1, column=pnl_r_col)

        if pnl_r_cell.value is not None and total_spent_amount_cell.value is not None:
            try:
                pnl_r_value = float(pnl_r_cell.value)
                total_spent_amount_value = float(total_spent_amount_cell.value)
                if pnl_r_value > total_spent_amount_value:
                    pnl_r_cell.fill = gold_fill
                else:
                    pnl_r_cell.fill = red_fill
            except ValueError:
                pass

        for row in worksheet.iter_rows(min_row=header_row + 1):
            delta_percentage_cell = row[delta_percentage_col - 1]
            delta_eth_cell = row[delta_eth_col - 1]
            dexscreener_cell = row[dexscreener_col - 1]
            number_buys_cell = row[number_buys_col - 1]
            number_sells_cell = row[number_sells_col - 1]
            token_symbol_cell = row[token_symbol_col - 1]
            outcome_cell = row[outcome_col - 1]
            incoming_cell = row[incoming_col - 1]

            try:
                if delta_percentage_cell.value is not None:
                    percentage_value = float(delta_percentage_cell.value)
                    if percentage_value == -100:
                        delta_percentage_cell.fill = brown_fill
                        delta_eth_cell.fill = red_fill
                    elif percentage_value > 0:
                        delta_percentage_cell.fill = green_fill
                        delta_eth_cell.fill = green_fill
                    elif percentage_value < 0:
                        delta_percentage_cell.fill = red_fill
                        delta_eth_cell.fill = red_fill

                    if dexscreener_cell.value:
                        original_url = dexscreener_cell.value
                        dexscreener_cell.value = "Dexscreener transaction"
                        dexscreener_cell.hyperlink = original_url
                        dexscreener_cell.font = Font(color="0000FF", underline="single")

            except ValueError:
                continue

        workbook.save(self.output_file_path)
        print(f'Excel file with conditional formatting has been saved to {self.output_file_path}')

    def generate_report(self):
        self.fetch_data()
        self.save_to_excel()
        self.apply_formatting()

if __name__ == "__main__":
    wallet_address = input("Enter the wallet address: ")
    report = WalletReport(wallet_address)
    report.generate_report()
